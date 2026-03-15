from pathlib import Path
import argparse
import openpyxl
import pandas as pd
import plotly.graph_objects as go

def load_pca_blocks(workbook_path: str, sheet_name: str = "ACP"):
    wb = openpyxl.load_workbook(workbook_path, data_only=True)
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Onglet '{sheet_name}' introuvable. Onglets disponibles : {wb.sheetnames}")
    ws = wb[sheet_name]

    variables = []
    for r in range(2, ws.max_row + 1):
        var = ws.cell(r, 1).value
        if var is None:
            break
        row = [ws.cell(r, c).value for c in (2, 3, 4)]
        if all(v is None for v in row):
            continue
        variables.append({"Variable": str(var), "PC1": float(row[0]), "PC2": float(row[1]), "PC3": float(row[2])})
    variables_df = pd.DataFrame(variables)
    variables_df["Norme"] = (variables_df[["PC1", "PC2", "PC3"]] ** 2).sum(axis=1).pow(0.5)

    subjects = []
    for r in range(2, ws.max_row + 1):
        sid = ws.cell(r, 15).value
        if sid is None:
            break
        row = [ws.cell(r, c).value for c in (16, 17, 18)]
        if all(v is None for v in row):
            continue
        sid = str(sid)
        subjects.append({
            "SubjectID": sid,
            "ID_court": sid[:10] + "…" if len(sid) > 12 else sid,
            "PC1": float(row[0]),
            "PC2": float(row[1]),
            "PC3": float(row[2]),
        })
    subjects_df = pd.DataFrame(subjects)
    subjects_df["Distance"] = (subjects_df[["PC1", "PC2", "PC3"]] ** 2).sum(axis=1).pow(0.5)

    explained = {}
    if "ACP_Results" in wb.sheetnames:
        wr = wb["ACP_Results"]
        for r in range(22, min(wr.max_row, 60) + 1):
            pc = wr.cell(r, 1).value
            expl = wr.cell(r, 3).value
            if isinstance(pc, str) and pc.startswith("PC") and pc[2:].isdigit() and isinstance(expl, (int, float)):
                explained[pc] = float(expl)

    return variables_df, subjects_df, explained

def make_html(workbook_path: str, output_html: str, sheet_name: str = "ACP"):
    import plotly.express as px

    variables_df, subjects_df, explained = load_pca_blocks(workbook_path, sheet_name=sheet_name)

    xlab = f"PC1 ({explained.get('PC1', 0):.1f}%)" if explained.get("PC1") is not None else "PC1"
    ylab = f"PC2 ({explained.get('PC2', 0):.1f}%)" if explained.get("PC2") is not None else "PC2"
    zlab = f"PC3 ({explained.get('PC3', 0):.1f}%)" if explained.get("PC3") is not None else "PC3"

    fig_subjects = px.scatter_3d(
        subjects_df, x="PC1", y="PC2", z="PC3", color="Distance",
        hover_name="ID_court",
        hover_data={"SubjectID": True, "PC1": ':.3f', "PC2": ':.3f', "PC3": ':.3f', "Distance": ':.3f'},
        title="Individus dans l'espace ACP"
    )
    fig_subjects.update_traces(marker={"size": 5})
    fig_subjects.update_layout(scene={"xaxis_title": xlab, "yaxis_title": ylab, "zaxis_title": zlab})

    fig_variables = go.Figure()
    for _, row in variables_df.iterrows():
        fig_variables.add_trace(go.Scatter3d(
            x=[0, row["PC1"]], y=[0, row["PC2"]], z=[0, row["PC3"]],
            mode="lines+markers+text",
            text=[None, row["Variable"]],
            textposition="top center",
            hovertemplate=f"<b>{row['Variable']}</b><br>PC1=%{{x:.3f}}<br>PC2=%{{y:.3f}}<br>PC3=%{{z:.3f}}<extra></extra>",
            showlegend=False
        ))
    fig_variables.update_layout(
        title="Variables / charges sur PC1-PC2-PC3",
        scene={"xaxis_title": xlab, "yaxis_title": ylab, "zaxis_title": zlab},
    )

    html = f"""
    <!DOCTYPE html>
    <html lang="fr">
    <head>
      <meta charset="utf-8">
      <title>Visualisation ACP</title>
      <style>
        body {{ font-family: Arial, sans-serif; margin: 24px; }}
        .card {{ background: #f7f7f9; border-radius: 12px; padding: 14px 18px; margin-bottom: 18px; }}
      </style>
    </head>
    <body>
      <h1>Visualisation ACP</h1>
      <div class="card">
        <p><b>Source :</b> {Path(workbook_path).name} — onglet <b>{sheet_name}</b></p>
        <p><b>Axes affichés :</b> {xlab}, {ylab}, {zlab}</p>
      </div>
      {fig_subjects.to_html(full_html=False, include_plotlyjs='cdn')}
      <br>
      {fig_variables.to_html(full_html=False, include_plotlyjs=False)}
    </body>
    </html>
    """
    Path(output_html).write_text(html, encoding="utf-8")

if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Créer une visualisation ACP 3D à partir d'un fichier Excel.")
    parser.add_argument("workbook", help="Chemin du fichier Excel")
    parser.add_argument("--sheet", default="ACP", help="Nom de l'onglet contenant les données ACP")
    parser.add_argument("--output", default="visualisation_acp.html", help="Fichier HTML de sortie")
    args = parser.parse_args()
    make_html(args.workbook, args.output, sheet_name=args.sheet)
